from tkinter import *
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import re
import os

# Fungsi untuk membuat file Excel baru
def create_new_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["Username", "Password", "Email", "Nama Lengkap"])  # Menambahkan header
    wb.save(path)
    return wb, ws

# Membuat atau membuka file Excel
file_path = 'user_data.xlsx'
if not os.path.exists(file_path):
    wb, ws = create_new_workbook(file_path)
else:
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            # Jika worksheet kosong, tambahkan header
            ws.append(["Username", "Password", "Email", "Nama Lengkap"])
            wb.save(file_path)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        wb, ws = create_new_workbook(file_path)

# Fungsi registrasi
def register(namalengkap, username, password, email):
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showerror("Error", "Email tidak valid. Silakan coba lagi.")
        return

    for row in ws.iter_rows(values_only=True):
        if row[0] == username:
            messagebox.showerror("Error", "Username sudah digunakan. Silakan coba lagi.")
            return

    ws.append([username, password, email, namalengkap])
    wb.save(file_path)
    messagebox.showinfo("Success", "Registrasi berhasil!")

# Fungsi login
def login(username, password):
    for row in ws.iter_rows(values_only=True):
        if row[0] == username and row[1] == password:
            messagebox.showinfo("Success", "Login berhasil!")
            halaman_utama()
            return

    messagebox.showerror("Error", "Username atau password salah.")

def halaman_login():
    global username_entry, password_entry, window1
    window1 = tk.Tk()
    window1.title("login")
    window1.geometry("1280x720")
    window1.configure(bg="black")
    window1.resizable(True, True)

    heading = tk.Label(window1, text="Sign In", fg="white", bg="black", font=("Montserrat", 23, "bold"))
    heading.place(x=825, y=170)

    username_label = tk.Label(window1, text="Username:", fg="white", bg="black", font=("Arial", 14))
    username_label.place(x=750, y=220)

    username_entry = tk.Entry(window1, width=25, fg="white", bg="black", font=("Arial", 11))
    username_entry.place(x=750, y=250)

    password_label = tk.Label(window1, text="Password:", fg="white", bg="black", font=("Arial", 14))
    password_label.place(x=750, y=290)

    password_entry = tk.Entry(window1, width=25, fg="white", bg="black", font=("Arial", 11), show="*")
    password_entry.place(x=750, y=320)

    img = Image.open("Kelompok-13/logo1.png")
    img = ImageTk.PhotoImage(img)
    label = tk.Label(window1, image=img, bd=0, highlightthickness=0)
    label.image = img
    label.place(x=150, y=55)

    def halaman_register():
        window1.destroy()
        window2 = tk.Tk()
        window2.title("register")
        window2.geometry("1280x720")
        window2.configure(bg="black")
        window2.resizable(True, True)

        heading = tk.Label(window2, text="Sign Up", fg="white", bg="black", font=("Montserrat", 23, "bold"))
        heading.place(x=825, y=100)

        new_namalengkap_label = tk.Label(window2, text="Nama Lengkap:", fg="white", bg="black", font=("Arial", 14))
        new_namalengkap_label.place(x=750, y=150)

        global new_namalengkap_entry
        new_namalengkap_entry = tk.Entry(window2, width=25, fg="white", bg="black", font=("Arial", 11))
        new_namalengkap_entry.place(x=750, y=180)

        new_username_label = tk.Label(window2, text="Username:", fg="white", bg="black", font=("Arial", 14))
        new_username_label.place(x=750, y=220)

        global new_username_entry
        new_username_entry = tk.Entry(window2, width=25, fg="white", bg="black", font=("Arial", 11))
        new_username_entry.place(x=750, y=250)

        new_password_label = tk.Label(window2, text="Password:", fg="white", bg="black", font=("Arial", 14))
        new_password_label.place(x=750, y=290)

        global new_password_entry
        new_password_entry = tk.Entry(window2, width=25, fg="white", bg="black", font=("Arial", 11), show="*")
        new_password_entry.place(x=750, y=320)

        new_email_label = tk.Label(window2, text="Email:", fg="white", bg="black", font=("Arial", 14))
        new_email_label.place(x=750, y=360)

        global new_email_entry
        new_email_entry = tk.Entry(window2, width=25, fg="white", bg="black", font=("Arial", 11))
        new_email_entry.place(x=750, y=390)

        img = Image.open("Kelompok-13/logo1.png")
        img = ImageTk.PhotoImage(img)
        label = tk.Label(window2, image=img, bd=0, highlightthickness=0)
        label.image = img
        label.place(x=105, y=55)

        def back_login():
            window2.destroy()
            halaman_login()

        register_button = tk.Button(window2, width=40, height=1, text="Register", fg="white", bg="black", command=process_registration)
        register_button.place(x=750, y=440)

        back_button = tk.Button(window2, width=40, height=1, text="Back", fg="white", bg="black", command=back_login)
        back_button.place(x=750, y=470)

        window2.mainloop()

    def check_login():
        username_input = username_entry.get()
        password_input = password_entry.get()
        login(username_input, password_input)

    login_button = tk.Button(window1, width=40, height=1, text="Login", fg="white", bg="black", command=check_login)
    login_button.place(x=750, y=370)

    new_info_label = tk.Label(window1, text="Don't have an account? Please register first.", fg="white", bg="black", font=("Arial", 8))
    new_info_label.place(x=750, y=405)

    signup_button = tk.Button(window1, width=40, height=1, text="Sign Up", fg="white", bg="black", command=halaman_register)
    signup_button.place(x=750, y=430)

    window1.mainloop()

def process_registration():
    new_namalengkap = new_namalengkap_entry.get()
    new_username = new_username_entry.get()
    new_password = new_password_entry.get()
    new_email = new_email_entry.get()
    register(new_namalengkap, new_username, new_password, new_email)

halaman_login()

# Jangan lupa untuk menutup workbook setelah selesai
wb.save(file_path)
wb.close()

       
